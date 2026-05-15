import asyncio
import os
import datetime
import random
import pandas as pd
from playwright.async_api import async_playwright
from core_analyzer import analyze_stock
from db_manager import save_report_to_db

# Finviz 週期對照表
PERIOD_MAPPING = {
    "1 month": {"p": "d", "r": "m1"},
    "3 month": {"p": "d", "r": "m3"},
    "6 month": {"p": "d", "r": "m6"},
    "year to date": {"p": "d", "r": "ytd"},
    "1 year": {"p": "d", "r": "y1"},
    "2 year": {"p": "w", "r": "y2"},
    "5 year": {"p": "m", "r": "y5"},
    "max": {"p": "m", "r": "max"}
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0"
]

async def internal_capture(url, output_path, selector=None, height=1500):
    """複刻 capture.py 的核心邏輯，確保參數一致"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        # 降級至 200 DPI (Token 節省約 50%+)
        scale_factor = 2.0 
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            viewport={"width": 1280, "height": int(height)},
            device_scale_factor=scale_factor
        )
        page = await context.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(5)
            await page.screenshot(path=output_path, full_page=False)
        finally:
            await browser.close()

def get_finviz_url(ticker, period_str):
    mapping = PERIOD_MAPPING.get(period_str.lower(), PERIOD_MAPPING["1 year"])
    return f"https://finviz.com/quote.ashx?t={ticker}&p={mapping['p']}&r={mapping['r']}"

def read_excel_safely(path, **kwargs):
    import shutil, tempfile
    if not os.path.exists(path): return None
    temp_path = os.path.join(tempfile.gettempdir(), f"temp_read_{os.path.basename(path)}")
    try:
        shutil.copy2(path, temp_path)
        return pd.read_excel(temp_path, **kwargs)
    finally:
        if os.path.exists(temp_path):
            try: os.remove(temp_path)
            except: pass

async def process_from_db(model_name="gemini-2.0-flash", target_period="1 year", tickers=None, progress_callback=None, max_concurrent=2):
    """
    從 MySQL 資料庫讀取監控清單並進行分析
    """
    from db_manager import get_watchlist_data
    
    # 取得監控清單 (包含 ticker, local_rule, default_period)
    watchlist = get_watchlist_data()
    
    process_list = []
    for item in watchlist:
        ticker = item['ticker']
        # 如果用戶指定了特定股票，則過濾
        if tickers and ticker not in tickers:
            continue
        
        # 優先順序：個股設定週期 > UI 傳入的全域週期
        final_period = item['default_period'] if item.get('default_period') else target_period
        
        process_list.append({
            "ticker": ticker,
            "period": final_period,
            "extra_rules": item['local_rule'] or "",
            "row_idx": 0 # 資料庫模式不需要 Excel 索引
        })

    if not process_list:
        return {}

    return await run_batch_core(process_list, model_name, capture_only=False, max_concurrent=max_concurrent, progress_callback=progress_callback)

async def run_batch_core(process_list, model_name, capture_only=False, max_concurrent=5, progress_callback=None):
    """
    核心批次處理邏輯 (被 Excel 和 DB 共享)
    """
    total = len(process_list)
    today = datetime.datetime.now().strftime("%Y%m%d")
    os.makedirs(f"captures/{today}", exist_ok=True)
    os.makedirs(f"reports/{today}", exist_ok=True)

    cache_name = None
    if not capture_only:
        if progress_callback: progress_callback(0, total, "SYSTEM", "running", "正在初始化 AI 全域快取...")
        try:
            from core_analyzer import create_global_cache
            # 模型映射
            cache_model = "gemini-2.5-flash" if "flash" in model_name else "gemini-2.5-pro"
            if "2.0" in model_name: cache_model = "gemini-2.0-flash"
            
            cache = create_global_cache(cache_model)
            cache_name = cache.name
            if progress_callback: progress_callback(0, total, "SYSTEM", "success", "AI 全域快取建立完成")
        except Exception as e:
            if progress_callback: progress_callback(0, total, "SYSTEM", "warning", f"快取建立失敗: {e}")

    semaphore = asyncio.Semaphore(max_concurrent)
    results = {}
    completed_count = 0

    async def wrapped_task(item):
        nonlocal completed_count
        async with semaphore:
            if progress_callback:
                progress_callback(completed_count, total, item["ticker"], "running", "開始擷取圖表與分析...")
            res = await process_single_stock_internal(item["ticker"], item["period"], item["extra_rules"], today, model_name, capture_only, item["row_idx"], cache_name)
            completed_count += 1
            if progress_callback:
                msg = res.get("summary", "完成") if res["status"] == "success" else res.get("message", "錯誤")
                progress_callback(completed_count, total, res["ticker"], res["status"], msg)
            return res

    tasks = [wrapped_task(item) for item in process_list]
    task_results_all = await asyncio.gather(*tasks)
    
    if cache_name:
        try:
            from google.generativeai import caching
            caching.CachedContent.get(cache_name).delete()
        except: pass

    for res in task_results_all:
        results[res["ticker"]] = res
    return results

async def process_from_excel(excel_path, model_name="gemini-2.5-flash", capture_only=False, max_concurrent=5, tickers_to_process=None, progress_callback=None):
    df = read_excel_safely(excel_path, header=None)
    if df is None: return {}

    process_list = []
    for index, row in df.iterrows():
        if index == 0: continue
        ticker = str(row.iloc[0]).strip().upper()
        if pd.isna(row.iloc[0]) or ticker in ["NAN", ""]: continue
        if tickers_to_process and ticker not in tickers_to_process: continue
        period = str(row.iloc[1]).strip().lower() if not pd.isna(row.iloc[1]) else "1 year"
        extra_rules = str(row.iloc[2]) if len(row) > 2 and not pd.isna(row.iloc[2]) else ""
        process_list.append({"ticker": ticker, "period": period, "extra_rules": extra_rules, "row_idx": index})

    results = await run_batch_core(process_list, model_name, capture_only, max_concurrent, progress_callback)
    if not capture_only:
        await write_results_to_excel(excel_path, list(results.values()))
    return results

async def process_single_stock_internal(ticker, period, extra_rules, date_str, model_name, capture_only, row_idx, cache_name=None):
    url = get_finviz_url(ticker, period)
    period_suffix = period.replace(' ', '_')
    image_path = f"captures/{date_str}/{ticker}_{period_suffix}.png"
    
    try:
        temp_image_path = f"captures/{date_str}/temp_{ticker}_{period_suffix}_{row_idx}.png"
        await internal_capture(url, temp_image_path)
        
        report_path = f"reports/{date_str}/{ticker}_{period_suffix}.md"
        capture_failed = False
        if os.path.exists(temp_image_path):
            if os.path.exists(image_path): os.remove(image_path)
            os.rename(temp_image_path, image_path)
        else:
            if os.path.exists(image_path): capture_failed = True
            else: raise Exception("截圖失敗且無舊圖")

        report_content = ""
        if not capture_only:
            report_content = await analyze_stock(ticker, image_path, model_name, extra_rules, cache_name)
            if capture_failed:
                file_time = datetime.datetime.fromtimestamp(os.path.getmtime(image_path)).strftime("%Y-%m-%d %H:%M:%S")
                report_content = f"> ⚠️ **注意**：今日截圖失敗，目前顯示的是舊有圖片（擷取時間：{file_time}）。\n\n" + report_content
            with open(report_path, "w", encoding="utf-8") as f: f.write(report_content)
        else:
            from db_manager import get_full_report
            existing_db = get_full_report(ticker, period, date_str)
            report_content = existing_db["report_content"] if existing_db else f"# {ticker} ({period}) 分析報告\n\n(等待 AI 分析中...)"
            with open(report_path, "w", encoding="utf-8") as f: f.write(report_content)

        try: save_report_to_db(ticker, period, report_content, image_path)
        except: pass
        
        return {"ticker": ticker, "status": "success", "row_idx": row_idx, "summary": report_content[:100].replace("\n", " ") + "..." if report_content else ""}
    except Exception as e:
        return {"ticker": ticker, "status": "error", "message": str(e), "row_idx": row_idx}

async def write_results_to_excel(excel_path, task_results):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        if ws.cell(row=1, column=4).value != "AI_Analysis":
            ws.cell(row=1, column=4).value = "AI_Analysis"
            ws.cell(row=1, column=5).value = "Last_Update"
        georgia_font = Font(name="Georgia", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for res in task_results:
            row = res["row_idx"] + 1
            curr_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            if res["status"] == "success":
                ws.cell(row=row, column=4).value = res["summary"]
                ws.cell(row=row, column=5).value = curr_time
            else:
                ws.cell(row=row, column=4).value = f"Error: {res.get('message', 'Unknown')}"
                ws.cell(row=row, column=5).value = curr_time
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = georgia_font
                cell.alignment = center_alignment
        wb.save(excel_path)
    except: pass
